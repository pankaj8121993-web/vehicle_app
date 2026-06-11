import io
from datetime import datetime, timezone, timedelta
from fastapi import APIRouter, Depends, HTTPException, Response
from database import db
from auth import require_user
from helpers import gather_expenses, get_lookup_maps

router = APIRouter(tags=["analytics"])


def today_str():
    return datetime.now(timezone.utc).strftime("%Y-%m-%d")


def month_start_str():
    return datetime.now(timezone.utc).strftime("%Y-%m-01")


async def latest_doc_expiries():
    """Latest expiry per (vehicle_id, doc_type)."""
    docs = await db.documents.find({"expiry_date": {"$ne": None}}, {"_id": 0}).to_list(5000)
    latest = {}
    for d in docs:
        key = (d["vehicle_id"], d["doc_type"])
        if key not in latest or (d.get("expiry_date") or "") > (latest[key].get("expiry_date") or ""):
            latest[key] = d
    return latest


async def compute_alerts():
    alerts = []
    today = today_str()
    in_30 = (datetime.now(timezone.utc) + timedelta(days=30)).strftime("%Y-%m-%d")
    vmap, _ = await get_lookup_maps()

    latest = await latest_doc_expiries()
    for (vid, dtype), d in latest.items():
        exp = d["expiry_date"]
        vnum = vmap.get(vid, "Unknown")
        if exp < today:
            alerts.append({"type": "document_expired", "severity": "danger",
                           "message": f"{dtype} EXPIRED for {vnum}", "vehicle_number": vnum, "due_date": exp})
        elif exp <= in_30:
            alerts.append({"type": "document_expiring", "severity": "warning",
                           "message": f"{dtype} expiring for {vnum}", "vehicle_number": vnum, "due_date": exp})

    drivers = await db.drivers.find({"license_expiry": {"$ne": None}}, {"_id": 0}).to_list(2000)
    for dr in drivers:
        exp = dr["license_expiry"]
        if exp < today:
            alerts.append({"type": "license_expired", "severity": "danger",
                           "message": f"License EXPIRED — {dr['name']}", "vehicle_number": "", "due_date": exp})
        elif exp <= in_30:
            alerts.append({"type": "license_expiring", "severity": "warning",
                           "message": f"License expiring — {dr['name']}", "vehicle_number": "", "due_date": exp})

    # Service due / overdue based on latest service per vehicle
    vehicles = await db.vehicles.find({}, {"_id": 0}).to_list(2000)
    in_7 = (datetime.now(timezone.utc) + timedelta(days=7)).strftime("%Y-%m-%d")
    for v in vehicles:
        latest_svc = await db.services.find({"vehicle_id": v["id"]}, {"_id": 0}).sort("date", -1).to_list(1)
        if latest_svc:
            s = latest_svc[0]
            due_date = s.get("next_due_date")
            due_km = s.get("next_due_km")
            odo = v.get("current_odometer") or 0
            if due_date and due_date < today:
                alerts.append({"type": "service_overdue", "severity": "danger",
                               "message": f"Service OVERDUE — {v['vehicle_number']}", "vehicle_number": v["vehicle_number"], "due_date": due_date})
            elif due_date and due_date <= in_7:
                alerts.append({"type": "service_due", "severity": "warning",
                               "message": f"Service due — {v['vehicle_number']}", "vehicle_number": v["vehicle_number"], "due_date": due_date})
            elif due_km and odo >= due_km:
                alerts.append({"type": "service_overdue", "severity": "danger",
                               "message": f"Service OVERDUE (by KM) — {v['vehicle_number']}", "vehicle_number": v["vehicle_number"], "due_date": None})
        if v.get("fastag_number") and (v.get("fastag_balance") or 0) < 200:
            alerts.append({"type": "fastag_low", "severity": "warning",
                           "message": f"Fastag balance low (₹{round(v.get('fastag_balance') or 0)}) — {v['vehicle_number']}",
                           "vehicle_number": v["vehicle_number"], "due_date": None})

    pending = await db.repairs.find({"status": "reported", "repair_type": "major"}, {"_id": 0}).to_list(500)
    for r in pending:
        alerts.append({"type": "repair_pending_approval", "severity": "warning",
                       "message": f"Repair pending approval — {vmap.get(r['vehicle_id'], '')}: {r.get('issue', '')}",
                       "vehicle_number": vmap.get(r["vehicle_id"], ""), "due_date": r.get("date")})

    order = {"danger": 0, "warning": 1}
    alerts.sort(key=lambda a: (order.get(a["severity"], 2), a.get("due_date") or "9999"))
    return alerts


@router.get("/alerts")
async def get_alerts(user=Depends(require_user)):
    return await compute_alerts()


@router.get("/dashboard")
async def dashboard(user=Depends(require_user)):
    today = today_str()
    in_30 = (datetime.now(timezone.utc) + timedelta(days=30)).strftime("%Y-%m-%d")
    month_start = month_start_str()
    vehicles = await db.vehicles.find({}, {"_id": 0}).to_list(2000)
    vmap = {v["id"]: v.get("vehicle_number", "") for v in vehicles}
    total_vehicles = len(vehicles)

    latest = await latest_doc_expiries()
    docs_expired = sum(1 for d in latest.values() if d["expiry_date"] < today)
    docs_expiring = sum(1 for d in latest.values() if today <= d["expiry_date"] <= in_30)
    licenses_expiring = await db.drivers.count_documents({"license_expiry": {"$gte": today, "$lte": in_30}})

    trips_today = await db.trips.find({"date": today}, {"_id": 0}).to_list(2000)
    running_today = len(set(t["vehicle_id"] for t in trips_today))
    active_trips = await db.trips.count_documents({"status": "ongoing"})
    under_repair_ids = set(r["vehicle_id"] for r in await db.repairs.find(
        {"status": {"$in": ["reported", "approved", "in_repair"]}}, {"_id": 0, "vehicle_id": 1}).to_list(2000))
    under_repair_ids |= set(v["id"] for v in vehicles if v.get("status") == "under_repair")
    vehicles_idle = max(total_vehicles - running_today - len(under_repair_ids), 0)

    month_fuel = await db.fuel_entries.find({"date": {"$gte": month_start}}, {"_id": 0}).to_list(5000)
    fuel_cost_month = round(sum(f.get("amount") or 0 for f in month_fuel), 2)
    all_mileages = [f["mileage"] for f in await db.fuel_entries.find({"mileage": {"$ne": None}}, {"_id": 0}).to_list(5000)]
    avg_mileage = round(sum(all_mileages) / len(all_mileages), 2) if all_mileages else None
    fuel_by_vehicle = {}
    for f in month_fuel:
        fuel_by_vehicle[f["vehicle_id"]] = fuel_by_vehicle.get(f["vehicle_id"], 0) + (f.get("amount") or 0)
    top_fuel = sorted(fuel_by_vehicle.items(), key=lambda x: -x[1])[:5]
    top_fuel = [{"vehicle_number": vmap.get(k, ""), "amount": round(v, 2)} for k, v in top_fuel]

    alerts = await compute_alerts()
    service_due = sum(1 for a in alerts if a["type"] == "service_due")
    service_overdue = sum(1 for a in alerts if a["type"] == "service_overdue")
    ninety_days_ago = (datetime.now(timezone.utc) - timedelta(days=90)).strftime("%Y-%m-%d")
    repair_counts = {}
    for r in await db.repairs.find({"date": {"$gte": ninety_days_ago}}, {"_id": 0}).to_list(5000):
        repair_counts[r["vehicle_id"]] = repair_counts.get(r["vehicle_id"], 0) + 1
    frequent_repairs = [{"vehicle_number": vmap.get(k, ""), "count": c} for k, c in sorted(repair_counts.items(), key=lambda x: -x[1]) if c >= 3][:5]

    month_expenses = await gather_expenses(start_date=month_start)
    monthly_cost = round(sum(r["amount"] for r in month_expenses), 2)
    month_trips = await db.trips.find({"date": {"$gte": month_start}}, {"_id": 0}).to_list(5000)
    month_km = sum(t.get("distance") or 0 for t in month_trips)
    cost_per_km = round(monthly_cost / month_km, 2) if month_km else None
    cost_by_vehicle = {}
    for r in month_expenses:
        cost_by_vehicle[r["vehicle_id"]] = cost_by_vehicle.get(r["vehicle_id"], 0) + r["amount"]
    top_cost = [{"vehicle_number": vmap.get(k, ""), "amount": round(v, 2)} for k, v in sorted(cost_by_vehicle.items(), key=lambda x: -x[1])[:5]]

    return {
        "compliance": {"total_vehicles": total_vehicles, "docs_expiring_30": docs_expiring,
                       "docs_expired": docs_expired, "licenses_expiring": licenses_expiring},
        "operations": {"running_today": running_today, "under_repair": len(under_repair_ids),
                       "idle": vehicles_idle, "active_trips": active_trips},
        "fuel": {"cost_this_month": fuel_cost_month, "avg_mileage": avg_mileage, "top_consumers": top_fuel},
        "maintenance": {"service_due": service_due, "service_overdue": service_overdue, "frequent_repairs": frequent_repairs},
        "financial": {"monthly_cost": monthly_cost, "cost_per_km": cost_per_km, "month_km": month_km, "top_cost_vehicles": top_cost},
        "alerts": alerts[:25],
    }


# ---------------- Reports ----------------
REPORT_LIST = [
    {"key": "trips", "name": "Trip Report"},
    {"key": "fuel", "name": "Fuel & Mileage Report"},
    {"key": "services", "name": "Service History Report"},
    {"key": "service_due", "name": "Service Due / Overdue List"},
    {"key": "repairs", "name": "Breakdown & Repair Report"},
    {"key": "documents", "name": "Document Compliance Report"},
    {"key": "expenses", "name": "Expense Ledger Report"},
    {"key": "expense_category", "name": "Category-wise Expense Summary"},
    {"key": "tyres", "name": "Tyre Performance Report"},
    {"key": "accidents", "name": "Accident Register Report"},
    {"key": "downtime", "name": "Downtime & Utilization Report"},
    {"key": "cost_per_km", "name": "Cost Per KM (Vehicle Ranking)"},
    {"key": "fuel_efficiency", "name": "Fuel Efficiency Ranking"},
]


def _q(vehicle_id, driver_id, start_date, end_date, date_field="date"):
    q = {}
    if vehicle_id:
        q["vehicle_id"] = vehicle_id
    if driver_id:
        q["driver_id"] = driver_id
    if start_date:
        q[date_field] = {"$gte": start_date}
    if end_date:
        q.setdefault(date_field, {})
        q[date_field]["$lte"] = end_date
    return q


async def build_report(key, vehicle_id=None, driver_id=None, start_date=None, end_date=None):
    vmap, dmap = await get_lookup_maps()
    today = today_str()

    if key == "trips":
        items = await db.trips.find(_q(vehicle_id, driver_id, start_date, end_date), {"_id": 0}).sort("date", -1).to_list(5000)
        cols = ["Date", "Vehicle", "Driver", "Origin", "Destination", "Opening KM", "Closing KM", "Distance (KM)", "Expenses (₹)", "Status"]
        rows = [[t.get("date"), vmap.get(t["vehicle_id"], ""), dmap.get(t.get("driver_id"), ""), t.get("origin"), t.get("destination"),
                 t.get("opening_km"), t.get("closing_km"), t.get("distance"),
                 (t.get("toll_expense") or 0) + (t.get("parking_expense") or 0) + (t.get("misc_expense") or 0), t.get("status")] for t in items]
        return cols, rows

    if key == "fuel":
        items = await db.fuel_entries.find(_q(vehicle_id, driver_id, start_date, end_date), {"_id": 0}).sort("date", -1).to_list(5000)
        cols = ["Date", "Vehicle", "Driver", "Odometer", "Quantity (L)", "Amount (₹)", "Mileage (KM/L)", "Cost/KM (₹)", "Station"]
        rows = [[f.get("date"), vmap.get(f["vehicle_id"], ""), dmap.get(f.get("driver_id"), ""), f.get("odometer"), f.get("quantity"),
                 f.get("amount"), f.get("mileage"), f.get("fuel_cost_per_km"), f.get("station")] for f in items]
        return cols, rows

    if key == "services":
        items = await db.services.find(_q(vehicle_id, None, start_date, end_date), {"_id": 0}).sort("date", -1).to_list(5000)
        cols = ["Date", "Vehicle", "Service Type", "Odometer", "Vendor", "Cost (₹)", "Next Due Date", "Next Due KM"]
        rows = [[s.get("date"), vmap.get(s["vehicle_id"], ""), s.get("service_type"), s.get("odometer"), s.get("vendor"),
                 s.get("cost"), s.get("next_due_date"), s.get("next_due_km")] for s in items]
        return cols, rows

    if key == "service_due":
        vehicles = await db.vehicles.find({}, {"_id": 0}).to_list(2000)
        cols = ["Vehicle", "Last Service", "Next Due Date", "Next Due KM", "Current Odometer", "Status"]
        rows = []
        for v in vehicles:
            latest = await db.services.find({"vehicle_id": v["id"]}, {"_id": 0}).sort("date", -1).to_list(1)
            if not latest:
                rows.append([v["vehicle_number"], "Never serviced", None, None, v.get("current_odometer"), "NO RECORD"])
                continue
            s = latest[0]
            odo = v.get("current_odometer") or 0
            status = "OK"
            if (s.get("next_due_date") and s["next_due_date"] < today) or (s.get("next_due_km") and odo >= s["next_due_km"]):
                status = "OVERDUE"
            elif s.get("next_due_date") and s["next_due_date"] <= (datetime.now(timezone.utc) + timedelta(days=15)).strftime("%Y-%m-%d"):
                status = "DUE SOON"
            rows.append([v["vehicle_number"], s.get("date"), s.get("next_due_date"), s.get("next_due_km"), odo, status])
        return cols, rows

    if key == "repairs":
        items = await db.repairs.find(_q(vehicle_id, None, start_date, end_date), {"_id": 0}).sort("date", -1).to_list(5000)
        cols = ["Date", "Vehicle", "Type", "Issue", "Vendor", "Cost (₹)", "Downtime (Days)", "Root Cause", "Status"]
        rows = [[r.get("date"), vmap.get(r["vehicle_id"], ""), r.get("repair_type"), r.get("issue"), r.get("vendor"),
                 r.get("cost"), r.get("downtime_days"), r.get("root_cause"), r.get("status")] for r in items]
        return cols, rows

    if key == "documents":
        q = {"vehicle_id": vehicle_id} if vehicle_id else {}
        items = await db.documents.find(q, {"_id": 0}).sort("expiry_date", 1).to_list(5000)
        cols = ["Vehicle", "Document", "Number", "Issue Date", "Expiry Date", "Status"]
        rows = []
        in_30 = (datetime.now(timezone.utc) + timedelta(days=30)).strftime("%Y-%m-%d")
        for d in items:
            exp = d.get("expiry_date")
            status = "VALID"
            if exp:
                status = "EXPIRED" if exp < today else ("EXPIRING SOON" if exp <= in_30 else "VALID")
            rows.append([vmap.get(d["vehicle_id"], ""), d.get("doc_type"), d.get("doc_number"), d.get("issue_date"), exp, status])
        return cols, rows

    if key == "expenses":
        ledger = await gather_expenses(vehicle_id=vehicle_id, start_date=start_date, end_date=end_date)
        cols = ["Date", "Vehicle", "Category", "Description", "Amount (₹)"]
        rows = [[r.get("date"), vmap.get(r.get("vehicle_id"), ""), r.get("category"), r.get("description"), r.get("amount")] for r in ledger]
        return cols, rows

    if key == "expense_category":
        ledger = await gather_expenses(vehicle_id=vehicle_id, start_date=start_date, end_date=end_date)
        agg = {}
        for r in ledger:
            agg[r["category"]] = agg.get(r["category"], 0) + r["amount"]
        cols = ["Category", "Total Amount (₹)"]
        rows = [[k, round(v, 2)] for k, v in sorted(agg.items(), key=lambda x: -x[1])]
        return cols, rows

    if key == "tyres":
        q = {"vehicle_id": vehicle_id} if vehicle_id else {}
        items = await db.tyres.find(q, {"_id": 0}).to_list(5000)
        cols = ["Tyre No", "Vehicle", "Brand", "Size", "Installed", "Install KM", "Removal KM", "Life (KM)", "Cost (₹)", "Punctures", "Status"]
        rows = []
        for t in items:
            punctures = await db.tyre_events.count_documents({"tyre_id": t["id"], "event_type": "puncture"})
            life = (t.get("removal_km") - t.get("installation_km")) if t.get("removal_km") and t.get("installation_km") else None
            rows.append([t.get("tyre_number"), vmap.get(t["vehicle_id"], ""), t.get("brand"), t.get("size"), t.get("installation_date"),
                         t.get("installation_km"), t.get("removal_km"), life, t.get("cost"), punctures, t.get("status")])
        return cols, rows

    if key == "accidents":
        items = await db.accidents.find(_q(vehicle_id, driver_id, start_date, end_date), {"_id": 0}).sort("date", -1).to_list(5000)
        cols = ["Date", "Vehicle", "Driver", "Location", "FIR", "Repair Cost (₹)", "Claim (₹)", "Settlement (₹)", "Claim Status"]
        rows = [[a.get("date"), vmap.get(a["vehicle_id"], ""), dmap.get(a.get("driver_id"), ""), a.get("location"), a.get("fir_number"),
                 a.get("repair_cost"), a.get("claim_amount"), a.get("settlement_amount"), a.get("claim_status")] for a in items]
        return cols, rows

    if key == "downtime":
        items = await db.downtimes.find(_q(vehicle_id, None, start_date, end_date, "start_date"), {"_id": 0}).sort("start_date", -1).to_list(5000)
        cols = ["Vehicle", "Reason", "Start", "End", "Days", "Status"]
        rows = [[vmap.get(d["vehicle_id"], ""), d.get("reason"), d.get("start_date"), d.get("end_date"), d.get("days"), d.get("status")] for d in items]
        return cols, rows

    if key == "cost_per_km":
        vehicles = await db.vehicles.find({}, {"_id": 0}).to_list(2000)
        cols = ["Vehicle", "Total KM", "Total Cost (₹)", "Cost/KM (₹)"]
        rows = []
        for v in vehicles:
            trips = await db.trips.find(_q(v["id"], None, start_date, end_date), {"_id": 0}).to_list(5000)
            km = sum(t.get("distance") or 0 for t in trips)
            cost = round(sum(r["amount"] for r in await gather_expenses(vehicle_id=v["id"], start_date=start_date, end_date=end_date)), 2)
            rows.append([v["vehicle_number"], km, cost, round(cost / km, 2) if km else None])
        rows.sort(key=lambda r: -(r[3] or 0))
        return cols, rows

    if key == "fuel_efficiency":
        vehicles = await db.vehicles.find({}, {"_id": 0}).to_list(2000)
        cols = ["Vehicle", "Fuel Entries", "Total Litres", "Total Fuel Cost (₹)", "Avg Mileage (KM/L)"]
        rows = []
        for v in vehicles:
            fuel = await db.fuel_entries.find(_q(v["id"], None, start_date, end_date), {"_id": 0}).to_list(5000)
            mileages = [f["mileage"] for f in fuel if f.get("mileage")]
            rows.append([v["vehicle_number"], len(fuel), round(sum(f.get("quantity") or 0 for f in fuel), 2),
                         round(sum(f.get("amount") or 0 for f in fuel), 2),
                         round(sum(mileages) / len(mileages), 2) if mileages else None])
        rows.sort(key=lambda r: -(r[4] or 0))
        return cols, rows

    raise HTTPException(status_code=404, detail="Unknown report")


@router.get("/reports")
async def list_reports(user=Depends(require_user)):
    return REPORT_LIST


@router.get("/reports/{key}")
async def get_report(key: str, vehicle_id: str = None, driver_id: str = None,
                     start_date: str = None, end_date: str = None, user=Depends(require_user)):
    cols, rows = await build_report(key, vehicle_id, driver_id, start_date, end_date)
    name = next((r["name"] for r in REPORT_LIST if r["key"] == key), key)
    return {"key": key, "name": name, "columns": cols, "rows": rows}


@router.get("/reports/{key}/export")
async def export_report(key: str, format: str = "excel", vehicle_id: str = None, driver_id: str = None,
                        start_date: str = None, end_date: str = None, user=Depends(require_user)):
    cols, rows = await build_report(key, vehicle_id, driver_id, start_date, end_date)
    name = next((r["name"] for r in REPORT_LIST if r["key"] == key), key)

    if format == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = name[:30]
        ws.append(cols)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        for r in rows:
            ws.append([("" if c is None else c) for c in r])
        for i, col in enumerate(cols, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(len(str(col)) + 4, 14)
        buf = io.BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{key}_report.xlsx"'},
        )

    if format == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=10 * mm, rightMargin=10 * mm, topMargin=12 * mm)
        styles = getSampleStyleSheet()
        elems = [Paragraph(f"Rajguru Foods — {name}", styles["Title"]),
                 Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y')}", styles["Normal"]),
                 Spacer(1, 8)]
        # PDF font lacks ₹ glyph; use Rs.
        safe_cols = [c.replace("₹", "Rs.") for c in cols]
        data = [safe_cols] + [[("" if c is None else str(c)) for c in r] for r in rows]
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elems.append(table)
        doc.build(elems)
        return Response(
            content=buf.getvalue(),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{key}_report.pdf"'},
        )

    raise HTTPException(status_code=400, detail="format must be excel or pdf")
