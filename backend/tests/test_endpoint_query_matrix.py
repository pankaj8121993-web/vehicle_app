"""UXR1-02: real-HTTP list/search/filter/pagination matrix on disposable MongoDB."""
import uuid
import io

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from passlib.hash import bcrypt

import database
import server
import session_security as ss
from conftest import realhttp_run as _run


MATRIX = [
    # path, collection, search field, default sort, supported sort, date field, domain filter
    ("vehicles", "vehicles", "vehicle_number", "vehicle_number", "make", None, ("status", "active")),
    ("drivers", "drivers", "name", "name", "employee_number", None, ("status", "active")),
    ("trips", "trips", "purpose", "date", "purpose", "date", ("vehicle_id", "matrix-vehicle")),
    ("expenses", "expenses", "description", "date", "amount", "date", ("category", "Fuel")),
    ("fuel", "fuel_entries", "fuel_station", "date", "fuel_station", "date", ("vehicle_id", "matrix-vehicle")),
    ("fastag", "fastag_transactions", "plaza_name", "date", "plaza_name", "date", ("vehicle_id", "matrix-vehicle")),
    ("repairs", "repairs", "description", "date", "ticket_number", "date", ("vehicle_id", "matrix-vehicle")),
    ("tyres", "tyres", "tyre_number", "installation_date", "brand", "installation_date", ("vehicle_id", "matrix-vehicle")),
    ("downtime", "downtimes", "reason", "start_date", "reason", "start_date", ("vehicle_id", "matrix-vehicle")),
    ("accidents", "accidents", "description", "date", "location", "date", ("vehicle_id", "matrix-vehicle")),
    ("documents", "documents", "doc_number", "expiry_date", "doc_number", "expiry_date", ("doc_type", "Insurance")),
    ("vendors", "vendors", "name", "name", "vendor_type", None, ("vendor_type", "Fuel")),
]


class Client:
    def __init__(self, http, token):
        self.http = http
        self.token = token

    async def req(self, method, path, **kwargs):
        kwargs.setdefault("headers", {})["Authorization"] = f"Bearer {self.token}"
        if method not in ss.SAFE_METHODS:
            csrf = self.http.cookies.get(ss.CSRF_COOKIE)
            if csrf:
                kwargs.setdefault("headers", {})[ss.CSRF_HEADER] = csrf
        return await self.http.request(method, path, **kwargs)


async def _login(username, password):
    http = AsyncClient(transport=ASGITransport(app=server.app), base_url="http://matrix")
    response = await http.post("/api/auth/login", json={"username": username, "password": password})
    assert response.status_code == 200, response.text
    return Client(http, response.json()["token"])


def _doc(collection, index, org_id, target=False):
    marker = "zzz-deep-page-needle" if target else f"ordinary-{index:02d}"
    day = f"2026-06-{(index % 28) + 1:02d}"
    doc = {
        "id": f"{org_id}-{collection}-{index}", "org_id": org_id,
        "created_at": f"{day}T00:00:00+00:00", "is_test_data": False,
        "status": "active" if index % 2 == 0 else "inactive",
        "vehicle_id": "matrix-vehicle" if index % 2 == 0 else "other-vehicle",
        "driver_id": "matrix-driver", "category": "Fuel" if index % 2 == 0 else "Other",
        "date": day, "start_date": day, "expiry_date": day, "installation_date": day,
        "amount": index + 0.5, "vehicle_number": marker, "make": f"Make-{index:02d}",
        "name": marker, "employee_number": f"EMP-{index:02d}", "purpose": marker,
        "description": marker, "fuel_station": marker, "plaza_name": marker,
        "ticket_number": f"TKT-{marker}", "tyre_number": marker, "brand": f"Brand-{index:02d}",
        "reason": marker, "location": marker, "doc_number": marker,
        "doc_type": "Insurance" if index % 2 == 0 else "Permit",
        "vendor_type": "Fuel" if index % 2 == 0 else "Repair",
        "is_active": index % 2 == 0, "claim_status": "reported",
    }
    # The target is intentionally last under ascending name-like defaults and
    # oldest under descending date defaults, so it is beyond page one.
    if target:
        doc.update({
            "date": "2025-01-01", "start_date": "2025-01-01",
            "expiry_date": "2025-01-01", "installation_date": "2025-01-01",
            "status": "active", "vehicle_id": "matrix-vehicle", "category": "Fuel",
            "doc_type": "Insurance", "vendor_type": "Fuel", "is_active": True,
        })
    return doc


@pytest.fixture(scope="module")
def matrix_env():
    async def build():
        await database.client.drop_database(database.raw_db.name)
        password = uuid.uuid4().hex + uuid.uuid4().hex
        org_a, org_b = "query-matrix-a", "query-matrix-b"
        await database.raw_db.organizations.insert_many([
            {"id": org_a, "legal_name": "Query Matrix A", "trade_name": "Query A"},
            {"id": org_b, "legal_name": "Query Matrix B", "trade_name": "Query B"},
        ])
        password_hash = bcrypt.hash(password)
        await database.raw_db.users.insert_many([
            {"id": "matrix-admin", "username": "matrix-admin", "password_hash": password_hash,
             "full_name": "Matrix Admin", "role": "org_admin", "org_id": org_a,
             "is_active": True, "must_change_password": False},
            {"id": "matrix-viewer", "username": "matrix-viewer", "password_hash": password_hash,
             "full_name": "Matrix Viewer", "role": "viewer", "org_id": org_a,
             "is_active": True, "must_change_password": False},
            {"id": "matrix-other", "username": "matrix-other", "password_hash": password_hash,
             "full_name": "Other Admin", "role": "org_admin", "org_id": org_b,
             "is_active": True, "must_change_password": False},
        ])
        collections = {entry[1] for entry in MATRIX}
        for collection in collections:
            own = [_doc(collection, i, org_a, target=i == 22) for i in range(23)]
            foreign = [_doc(collection, i, org_b, target=i == 22) for i in range(7)]
            await database.raw_db[collection].insert_many(own + foreign)
        return {
            "admin": await _login("matrix-admin", password),
            "viewer": await _login("matrix-viewer", password),
            "other": await _login("matrix-other", password),
        }

    env = _run(build())
    yield env

    async def teardown():
        for client in env.values():
            await client.http.aclose()
        await database.client.drop_database(database.raw_db.name)

    _run(teardown())


def _get(client, path, params=None):
    response = _run(client.req("GET", f"/api/{path}", params=params or {}))
    assert response.status_code == 200, f"{path}: {response.status_code} {response.text[:300]}"
    return response.json()


@pytest.mark.parametrize(
    "path,collection,search_field,default_sort,supported_sort,date_field,domain_filter", MATRIX
)
def test_complete_query_contract(
    matrix_env, path, collection, search_field, default_sort,
    supported_sort, date_field, domain_filter,
):
    admin = matrix_env["admin"]
    first = _get(admin, path, {"page": 1, "page_size": 5})
    assert first["page"] == 1 and first["page_size"] == 5
    assert first["total"] == 23 and first["total_pages"] == 5
    assert len(first["items"]) == 5
    assert not any(item.get(search_field) == "zzz-deep-page-needle" for item in first["items"])

    second = _get(admin, path, {"page": 2, "page_size": 5})
    assert second["page"] == 2 and len(second["items"]) == 5
    last = _get(admin, path, {"page": 5, "page_size": 5})
    assert len(last["items"]) == 3
    empty = _get(admin, path, {"page": 99, "page_size": 5})
    assert empty["items"] == [] and empty["total"] == 23

    bounded = _get(admin, path, {"page": -5, "page_size": 9999})
    assert bounded["page"] == 1 and bounded["page_size"] == 200
    invalid = _get(admin, path, {"page": "invalid", "page_size": "invalid"})
    assert invalid["page"] == 1 and invalid["page_size"] == 25

    asc = _get(admin, path, {"sort_by": supported_sort, "sort_dir": "asc", "page_size": 200})
    desc = _get(admin, path, {"sort_by": supported_sort, "sort_dir": "desc", "page_size": 200})
    asc_values = [item.get(supported_sort) for item in asc["items"]]
    desc_values = [item.get(supported_sort) for item in desc["items"]]
    assert asc_values == sorted(asc_values)
    assert desc_values == sorted(desc_values, reverse=True)
    unsupported = _get(admin, path, {"sort_by": "$where", "sort_dir": "asc", "page_size": 5})
    assert unsupported["total"] == 23 and len(unsupported["items"]) == 5

    searched = _get(admin, path, {"search": "deep-page-needle", "page_size": 5})
    assert searched["total"] == 1
    assert searched["items"][0][search_field] == "zzz-deep-page-needle"
    assert _get(admin, path, {"search": "does-not-exist"})["items"] == []

    key, value = domain_filter
    filtered = _get(admin, path, {key: value, "page_size": 200})
    assert filtered["total"] == 12
    combined = _get(admin, path, {key: value, "search": "deep-page-needle"})
    assert combined["total"] == 1
    status = _get(admin, path, {"status": "active", "page_size": 200})
    if path not in {"vendors"}:
        assert status["total"] == 12
    if date_field:
        dated = _get(admin, path, {"start_date": "2026-06-10", "end_date": "2026-06-20", "page_size": 200})
        assert 0 < dated["total"] < 23


@pytest.mark.parametrize("path", [entry[0] for entry in MATRIX])
def test_tenant_and_permission_enforcement(matrix_env, path):
    own = _get(matrix_env["admin"], path, {"page_size": 200})
    other = _get(matrix_env["other"], path, {"page_size": 200})
    assert own["total"] == 23 and other["total"] == 7
    assert all(item["org_id"] == "query-matrix-a" for item in own["items"])
    assert all(item["org_id"] == "query-matrix-b" for item in other["items"])

    anonymous = _run(AsyncClient(
        transport=ASGITransport(app=server.app), base_url="http://anonymous"
    ).get(f"/api/{path}"))
    assert anonymous.status_code == 401
    viewer = _get(matrix_env["viewer"], path, {"page": 1, "page_size": 1})
    assert viewer["total"] == 23


def test_exceptions_are_tenant_scoped_filtered_and_permission_enforced(matrix_env):
    admin = matrix_env["admin"]
    other = matrix_env["other"]
    own = _get(admin, "exceptions", {"doc_horizon_days": 10000})
    foreign = _get(other, "exceptions", {"doc_horizon_days": 10000})
    assert own["total"] > foreign["total"] > 0
    category = own["items"][0]["category"]
    filtered = _get(admin, "exceptions", {"doc_horizon_days": 10000, "category": category})
    assert filtered["total"] > 0
    assert all(item["category"] == category for item in filtered["items"])
    invalid = _run(admin.req("GET", "/api/exceptions", params={"doc_horizon_days": -1}))
    assert invalid.status_code == 422
    denied = _run(matrix_env["viewer"].req(
        "POST", f"/api/exceptions/{own['items'][0]['id']}/acknowledge", json={}
    ))
    assert denied.status_code == 403


def test_export_filter_parity_full_result_and_safe_content(matrix_env):
    admin = matrix_env["admin"]
    params = {
        "vehicle_id": "matrix-vehicle", "start_date": "2026-06-10",
        "end_date": "2026-06-20",
    }
    report = _get(admin, "reports/trips", params)
    assert len(report["rows"]) > 1
    excel = _run(admin.req(
        "GET", "/api/reports/trips/export", params={**params, "format": "excel"}
    ))
    assert excel.status_code == 200
    assert excel.headers["x-export-row-count"] == str(len(report["rows"]))
    assert excel.headers["x-export-row-limit"] == "5000"
    assert excel.headers["x-export-truncated"] == "false"
    assert excel.headers["content-disposition"] == 'attachment; filename="trips_report.xlsx"'
    sheet = load_workbook(io.BytesIO(excel.content), read_only=True).active
    values = list(sheet.values)
    assert list(values[0]) == report["columns"]
    assert len(values) - 1 == len(report["rows"])
    flattened = "\n".join(str(cell) for row in values for cell in row)
    assert "query-matrix-b" not in flattened

    pdf = _run(admin.req(
        "GET", "/api/reports/trips/export", params={**params, "format": "pdf"}
    ))
    assert pdf.status_code == 200 and pdf.content.startswith(b"%PDF-")
    assert pdf.headers["x-export-row-count"] == str(len(report["rows"]))
    assert pdf.headers["content-disposition"] == 'attachment; filename="trips_report.pdf"'

    empty = _run(admin.req(
        "GET", "/api/reports/trips/export",
        params={"vehicle_id": "missing", "format": "excel"},
    ))
    empty_sheet = load_workbook(io.BytesIO(empty.content), read_only=True).active
    assert len(list(empty_sheet.values)) == 1
    assert empty.headers["x-export-row-count"] == "0"
    invalid = _run(admin.req("GET", "/api/reports/trips/export", params={"format": "csv"}))
    assert invalid.status_code == 400
    anonymous = _run(AsyncClient(
        transport=ASGITransport(app=server.app), base_url="http://anonymous-export"
    ).get("/api/reports/trips/export"))
    assert anonymous.status_code == 401


@pytest.mark.parametrize("key", [
    "trips", "fuel", "services", "service_due", "repairs", "documents",
    "expenses", "expense_category", "tyres", "accidents", "downtime",
    "cost_per_km", "fuel_efficiency", "greasing", "greasing_due",
])
def test_every_synchronous_export_is_authenticated_and_parseable(matrix_env, key):
    response = _run(matrix_env["admin"].req(
        "GET", f"/api/reports/{key}/export", params={"format": "excel"}
    ))
    assert response.status_code == 200, f"{key}: {response.text[:200]}"
    assert response.headers["content-disposition"] == f'attachment; filename="{key}_report.xlsx"'
    workbook = load_workbook(io.BytesIO(response.content), read_only=True)
    assert workbook.active.max_row >= 1
    assert int(response.headers["x-export-row-count"]) == workbook.active.max_row - 1


def test_spreadsheet_formula_fields_are_escaped():
    from routes_analytics import _safe_excel_cell
    assert _safe_excel_cell("=HYPERLINK('bad')").startswith("'")
    assert _safe_excel_cell("+cmd").startswith("'")
    assert _safe_excel_cell("ordinary") == "ordinary"
